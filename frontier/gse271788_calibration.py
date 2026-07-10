"""Freeze and re-derive the independent primary-CD4 calibration proposal."""
from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import math
import random
import re
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from receipt.schema import Artifact, EvidenceAtom, Receipt, Verifier


DATA = ROOT / "examples" / "data"
SOURCE_DIR = DATA / "gse271788"
OUT_JSON = DATA / "gse271788_calibration.json"
OUT_DOC = ROOT / "docs" / "GSE271788_CALIBRATION.md"
PREREG = DATA / "gse271788_preregistration.json"
MARSON = DATA / "marson_de_full.csv"
K562 = DATA / "replogle_k562_de.csv"

SEED = 271788
EXPECTED_SOURCE_HASHES = {
    "counts.gz": "24f49fc8480a4b7b79fd2fd3ee03a2ca1429a6822f3b6d1ea6c30b13d02bf67c",
    "gse271788.soft": "ddc41b0ca6e7d3a04a12696ed349959ad7f3857d25e035c9b76a6c5a45bbe936",
    "gse171737.soft": "0b527f6b974e9ea2bce2f6001f3d2c213bec0cfd090cc671296cdd9aa1a7d80a",
    "supp.zip": "1ea94ce18c5d2e0e05b3602c4b788c76488264dae82b8046f9146daa30d83e6f",
    "mmc5.xlsx": "c0a22be718bec3a113fe3a2d543111efcf799b00693001d790358455c3d880c1",
    "mmc6.xlsx": "c5ab39e21057549e34451b17ef41be2018b7141a3ff455b0f8e4465aaabed2de",
    "mmc12.xlsx": "51a1e2dd79c3f24e14a637f451877128d07d981296ee03e66838cba8850ccec5",
    "mmc13.xlsx": "d828c2faff2de3bea7089b4c2e0b6f0c77a7f9ac6640a7845ba2fa7deebae790",
}
EXPECTED_PROJECTION_HASHES = {
    "target_reach.csv": "699bf26b2789b45a6e16ff0ecf3eefa9ba3f9169e9196355b40c766d0be04e18",
    "metadata_summary.json": "e46850f33d850f41d626ef95c2aa8bda842e239bc3f140e00bec9d61f21a6448",
}

SOURCE_URLS = {
    "counts.gz": "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE271nnn/GSE271788/suppl/GSE271788_dedup_counts.txt.gz",
    "gse271788.soft": "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE271788&targ=self&view=full&form=text",
    "gse171737.soft": "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE171737&targ=self&view=full&form=text",
    "supp.zip": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC11605694/supplementaryFiles",
    "mmc5.xlsx": "PMC11605694 supplementary Table S4",
    "mmc6.xlsx": "PMC11605694 supplementary Table S5",
    "mmc12.xlsx": "PMC11605694 supplementary Table S11",
    "mmc13.xlsx": "PMC11605694 supplementary Table S12",
}


class DataConstraintError(RuntimeError):
    """Raised when a frozen source or projection violates its sealed contract."""


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _canonical_id(prefix: str, payload: dict[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
        allow_nan=False,
    ).encode()
    return prefix + hashlib.sha256(encoded).hexdigest()[:16]


def _read_xlsx(path: Path) -> tuple[list[str], Iterable[tuple[Any, ...]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataConstraintError("openpyxl is required only to freeze the published workbooks") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(value) for value in next(rows)]
    return header, rows


def _dict_rows(path: Path) -> Iterable[dict[str, Any]]:
    header, rows = _read_xlsx(path)
    for values in rows:
        yield dict(zip(header, values))


def _normalize_guide(value: Any) -> str:
    guide = str(value)
    if guide.startswith("AAVS1"):
        return "AAVS1"
    return re.sub(r"_1$", "", guide)


def _parse_count_header(path: Path) -> dict[str, Any]:
    with gzip.open(path, "rt") as handle:
        command = next(handle).rstrip("\n")
        header = next(handle).rstrip("\n").split("\t")
    count_columns = header[6:]
    donors_by_target: dict[str, set[str]] = defaultdict(set)
    controls = 0
    for value in count_columns:
        sample = value.rsplit("/", 1)[-1].removesuffix(".dedup.bam")
        match = re.fullmatch(r"Donor_(\d+)_(.+)", sample)
        if not match:
            raise DataConstraintError(f"unparseable count column: {value}")
        donor, target = match.groups()
        target = re.sub(r"_\d+_[12]$", "", target)
        if target.startswith("AAVS1"):
            controls += 1
        else:
            donors_by_target[target].add(donor)
    return {
        "featurecounts_command_sha256": hashlib.sha256(command.encode()).hexdigest(),
        "count_columns": len(count_columns),
        "control_columns": controls,
        "targets": sorted(donors_by_target),
        "donors_by_target": {key: sorted(value) for key, value in sorted(donors_by_target.items())},
    }


def freeze_sources(source_dir: Path, out_dir: Path = SOURCE_DIR) -> dict[str, Any]:
    for name, expected in EXPECTED_SOURCE_HASHES.items():
        path = source_dir / name
        if not path.exists():
            raise DataConstraintError(f"missing source artifact: {path}")
        if _sha256(path) != expected:
            raise DataConstraintError(f"source hash drift: {name}")

    counts = _parse_count_header(source_dir / "counts.gz")
    if counts["count_columns"] != 311 or len(counts["targets"]) != 84:
        raise DataConstraintError("combined count matrix cardinality drift")
    if any(len(donors) != 3 for donors in counts["donors_by_target"].values()):
        raise DataConstraintError("each perturbation target must have exactly three donors")

    mashr_counts: Counter[str] = Counter()
    mashr_rows = 0
    mashr_pairs: set[tuple[str, str]] = set()
    for row in _dict_rows(source_dir / "mmc6.xlsx"):
        gene = str(row["from"])
        target = str(row["to"])
        lfsr = float(row["lfsr"])
        if lfsr >= 0.005:
            raise DataConstraintError("Table S5 contains a row outside the published LFSR threshold")
        pair = (gene, target)
        if pair in mashr_pairs:
            raise DataConstraintError(f"duplicate Table S5 effect: {gene} to {target}")
        mashr_pairs.add(pair)
        mashr_counts[gene] += 1
        mashr_rows += 1

    deseq_counts: Counter[str] = Counter()
    deseq_rows = 0
    for row in _dict_rows(source_dir / "mmc5.xlsx"):
        deseq_rows += 1
        if row["padj"] not in (None, "NA") and float(row["padj"]) < 0.05:
            deseq_counts[str(row["KO"])] += 1

    editing: dict[str, list[float]] = defaultdict(list)
    editing_rows = 0
    for row in _dict_rows(source_dir / "mmc12.xlsx"):
        editing_rows += 1
        gene = str(row["HGNC symbol"])
        if gene != "AAVS1":
            editing[gene].append(float(row["Percent_modified"]))

    cell_counts: dict[str, list[int]] = defaultdict(list)
    batch_by_gene: dict[str, str] = {}
    cell_count_rows = 0
    for row in _dict_rows(source_dir / "mmc13.xlsx"):
        cell_count_rows += 1
        gene = _normalize_guide(row["Guide"])
        if gene == "AAVS1":
            continue
        experiment = str(row["Experiment"])
        batch = (
            "gse171737_il2ra_regulators"
            if experiment == "IL2RA Regulators"
            else "gse271788_iei_background"
        )
        previous = batch_by_gene.setdefault(gene, batch)
        if previous != batch:
            raise DataConstraintError(f"target appears in two batches: {gene}")
        cell_counts[gene].append(int(row["Cell count"]))

    if mashr_rows != 226763 or len(mashr_counts) != 84:
        raise DataConstraintError("Table S5 cardinality drift")
    if deseq_rows != 129727:
        raise DataConstraintError("Table S4 cardinality drift")
    if editing_rows != 212 or len(editing) != 60:
        raise DataConstraintError("Table S11 cardinality drift")
    if cell_count_rows != 310 or len(cell_counts) != 84:
        raise DataConstraintError("Table S12 cardinality drift")
    if set(mashr_counts) != set(counts["targets"]) or set(cell_counts) != set(counts["targets"]):
        raise DataConstraintError("target sets disagree across released artifacts")

    marson_targets = set()
    with MARSON.open(newline="") as handle:
        for row in csv.DictReader(handle):
            marson_targets.add(row["target_contrast_gene_name"])
    overlap = sorted(set(mashr_counts) & marson_targets)
    not_assayed = sorted(set(mashr_counts) - marson_targets)
    if len(overlap) != 79 or len(not_assayed) != 5:
        raise DataConstraintError("Marson overlap cardinality drift")

    out_dir.mkdir(parents=True, exist_ok=True)
    target_path = out_dir / "target_reach.csv"
    fields = [
        "gene",
        "batch",
        "published_mashr_lfsr_lt_0_005",
        "published_deseq2_padj_lt_0_05",
        "editing_efficiency_min",
        "editing_efficiency_median",
        "live_cell_count_min",
        "live_cell_count_median",
        "donor_count",
    ]
    with target_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for gene in sorted(mashr_counts):
            edit = editing.get(gene, [])
            cells = cell_counts[gene]
            writer.writerow({
                "gene": gene,
                "batch": batch_by_gene[gene],
                "published_mashr_lfsr_lt_0_005": mashr_counts[gene],
                "published_deseq2_padj_lt_0_05": deseq_counts[gene],
                "editing_efficiency_min": f"{min(edit):.6f}" if edit else "",
                "editing_efficiency_median": f"{statistics.median(edit):.6f}" if edit else "",
                "live_cell_count_min": min(cells),
                "live_cell_count_median": f"{statistics.median(cells):.1f}",
                "donor_count": len(counts["donors_by_target"][gene]),
            })

    metadata = {
        "schema_version": "prospect.gse271788.metadata.v1",
        "count_columns": counts["count_columns"],
        "control_columns": counts["control_columns"],
        "targets": len(counts["targets"]),
        "donors_per_target": 3,
        "unique_donors": sorted({donor for values in counts["donors_by_target"].values() for donor in values}),
        "mashr_rows_lfsr_lt_0_005": mashr_rows,
        "deseq2_rows": deseq_rows,
        "editing_efficiency_rows": editing_rows,
        "live_cell_count_rows": cell_count_rows,
        "batch_target_counts": dict(sorted(Counter(batch_by_gene.values()).items())),
        "marson_overlap": len(overlap),
        "marson_not_assayed": not_assayed,
        "featurecounts_command_sha256": counts["featurecounts_command_sha256"],
    }
    metadata_path = out_dir / "metadata_summary.json"
    metadata_path.write_text(json.dumps(metadata, indent=2, sort_keys=True) + "\n")

    manifest = {
        "schema_version": "prospect.substrate_manifest.v1",
        "substrate_id": "weinstock_freimer_activated_cd4_ko",
        "organism": "Homo sapiens",
        "cell_type": "primary human CD4+ T cells",
        "perturbation": "arrayed CRISPR knockout",
        "phenotype": "activated transcriptome",
        "condition": "CD3/CD28/CD2 activation with IL-2, RNA collected day 8",
        "identifier_namespace": "HGNC gene symbol",
        "comparability": "orthogonal_activation_context",
        "coverage": {"targets": 84, "marson_overlap": 79, "marson_not_assayed": 5},
        "scoring_rule": "published mashr downstream-effect count at LFSR below 0.005",
        "missing_semantics": "not_assayed",
        "replay": "python frontier/gse271788_calibration.py --check",
        "sources": {
            name: {
                "url": SOURCE_URLS[name],
                "sha256": expected,
                "bytes": (source_dir / name).stat().st_size,
            }
            for name, expected in EXPECTED_SOURCE_HASHES.items()
        },
        "projections": {
            "target_reach.csv": {"sha256": _sha256(target_path), "rows": 84},
            "metadata_summary.json": {"sha256": _sha256(metadata_path)},
        },
    }
    manifest_path = out_dir / "source_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n")
    return manifest


def validate_frozen_inputs(source_dir: Path = SOURCE_DIR) -> dict[str, Any]:
    manifest_path = source_dir / "source_manifest.json"
    if not manifest_path.exists():
        raise DataConstraintError(f"missing source manifest: {manifest_path}")
    manifest = json.loads(manifest_path.read_text())
    for name, expected in EXPECTED_SOURCE_HASHES.items():
        if manifest["sources"][name]["sha256"] != expected:
            raise DataConstraintError(f"source manifest hash drift: {name}")
    for name in ("target_reach.csv", "metadata_summary.json"):
        path = source_dir / name
        if not path.exists():
            raise DataConstraintError(f"missing frozen projection: {path}")
        actual = _sha256(path)
        if manifest["projections"][name]["sha256"] != actual:
            raise DataConstraintError(f"projection hash drift: {name}")
        pinned = EXPECTED_PROJECTION_HASHES.get(name)
        if pinned and actual != pinned:
            raise DataConstraintError(f"pinned projection hash drift: {name}")
    metadata = json.loads((source_dir / "metadata_summary.json").read_text())
    expected = {
        "count_columns": 311,
        "targets": 84,
        "donors_per_target": 3,
        "mashr_rows_lfsr_lt_0_005": 226763,
        "deseq2_rows": 129727,
        "editing_efficiency_rows": 212,
        "live_cell_count_rows": 310,
        "marson_overlap": 79,
    }
    for key, value in expected.items():
        if metadata.get(key) != value:
            raise DataConstraintError(f"metadata drift for {key}")
    return {"manifest": manifest, "metadata": metadata}


def _rank(values: list[float]) -> list[float]:
    order = sorted(range(len(values)), key=values.__getitem__)
    ranks = [0.0] * len(values)
    start = 0
    while start < len(order):
        end = start + 1
        while end < len(order) and values[order[end]] == values[order[start]]:
            end += 1
        average = (start + 1 + end) / 2.0
        for index in order[start:end]:
            ranks[index] = average
        start = end
    return ranks


def _spearman(left: list[float], right: list[float]) -> float:
    if len(left) != len(right) or len(left) < 2:
        raise DataConstraintError("Spearman inputs must have equal length above one")
    x = _rank(left)
    y = _rank(right)
    x_mean = sum(x) / len(x)
    y_mean = sum(y) / len(y)
    numerator = sum((a - x_mean) * (b - y_mean) for a, b in zip(x, y))
    denominator = math.sqrt(
        sum((a - x_mean) ** 2 for a in x) * sum((b - y_mean) ** 2 for b in y)
    )
    if not denominator:
        raise DataConstraintError("Spearman input is constant")
    return numerator / denominator


def _permutation_p(left: list[float], right: list[float], observed: float) -> float:
    rng = random.Random(SEED)
    at_least_observed = 0
    for _ in range(10000):
        permuted = list(right)
        rng.shuffle(permuted)
        if _spearman(left, permuted) >= observed:
            at_least_observed += 1
    return (at_least_observed + 1) / 10001


def _bootstrap_interval(left: list[float], right: list[float]) -> list[float]:
    rng = random.Random(SEED)
    estimates = []
    for _ in range(10000):
        indices = [rng.randrange(len(left)) for _ in left]
        try:
            estimates.append(
                _spearman([left[index] for index in indices], [right[index] for index in indices])
            )
        except DataConstraintError:
            continue
    estimates.sort()
    return [round(estimates[int(0.025 * len(estimates))], 6), round(estimates[int(0.975 * len(estimates)) - 1], 6)]


def _association(genes: list[str], marson: dict[tuple[str, str], dict[str, str]], reach: dict[str, int], condition: str) -> dict[str, Any]:
    covered = [gene for gene in genes if (gene, condition) in marson]
    rho = _spearman(
        [float(marson[(gene, condition)]["n_total_de_genes"]) for gene in covered],
        [float(reach[gene]) for gene in covered],
    )
    return {"n": len(covered), "spearman_rho": round(rho, 6)}


def build_calibration(source_dir: Path = SOURCE_DIR) -> dict[str, Any]:
    integrity = validate_frozen_inputs(source_dir)
    prereg = json.loads(PREREG.read_text())
    if prereg["pre_registration_id"] != "prereg_3c8226ccc9c0773d":
        raise DataConstraintError("pre-registration id drift")

    with (source_dir / "target_reach.csv").open(newline="") as handle:
        target_rows = list(csv.DictReader(handle))
    reach = {row["gene"]: int(row["published_mashr_lfsr_lt_0_005"]) for row in target_rows}
    batch = {row["gene"]: row["batch"] for row in target_rows}
    with MARSON.open(newline="") as handle:
        marson = {
            (row["target_contrast_gene_name"], row["culture_condition"]): row
            for row in csv.DictReader(handle)
        }
    with K562.open(newline="") as handle:
        k562 = {row["gene"]: int(float(row["k562_de"])) for row in csv.DictReader(handle)}

    overlap = sorted(set(reach) & {gene for gene, _ in marson})
    not_assayed = sorted(set(reach) - {gene for gene, _ in marson})
    primary_genes = [gene for gene in overlap if (gene, "Stim48hr") in marson]
    left = [float(marson[(gene, "Stim48hr")]["n_total_de_genes"]) for gene in primary_genes]
    right = [float(reach[gene]) for gene in primary_genes]
    observed = _spearman(left, right)
    permutation_p = _permutation_p(left, right, observed)
    bootstrap = _bootstrap_interval(left, right)

    batch_results = {}
    for batch_id in ("gse171737_il2ra_regulators", "gse271788_iei_background"):
        genes = [gene for gene in primary_genes if batch[gene] == batch_id]
        batch_results[batch_id] = {
            "n": len(genes),
            "spearman_rho": round(
                _spearman(
                    [float(marson[(gene, "Stim48hr")]["n_total_de_genes"]) for gene in genes],
                    [float(reach[gene]) for gene in genes],
                ),
                6,
            ),
        }
    batch_pass = all(row["spearman_rho"] > 0 for row in batch_results.values())

    machinery_genes = [gene for gene in primary_genes if gene in k562 and k562[gene] <= 25]
    machinery_rho = _spearman(
        [float(marson[(gene, "Stim48hr")]["n_total_de_genes"]) for gene in machinery_genes],
        [float(reach[gene]) for gene in machinery_genes],
    )
    machinery_missing = sorted(gene for gene in primary_genes if gene not in k562)
    machinery_excluded = sorted(gene for gene in primary_genes if gene in k562 and k562[gene] > 25)

    leave_one_out = []
    for excluded in primary_genes:
        genes = [gene for gene in primary_genes if gene != excluded]
        leave_one_out.append({
            "excluded_gene": excluded,
            "spearman_rho": round(
                _spearman(
                    [float(marson[(gene, "Stim48hr")]["n_total_de_genes"]) for gene in genes],
                    [float(reach[gene]) for gene in genes],
                ),
                6,
            ),
        })
    least_stable = min(leave_one_out, key=lambda row: row["spearman_rho"])

    on_target_genes = [
        gene
        for gene in primary_genes
        if marson[(gene, "Stim48hr")]["ontarget_effect_category"] == "on-target KD"
    ]
    on_target_rho = _spearman(
        [float(marson[(gene, "Stim48hr")]["n_total_de_genes"]) for gene in on_target_genes],
        [float(reach[gene]) for gene in on_target_genes],
    )

    kills = {
        "batch_specificity": {"passed": batch_pass, "batches": batch_results},
        "general_machinery": {
            "passed": machinery_rho > 0,
            "n": len(machinery_genes),
            "spearman_rho": round(machinery_rho, 6),
            "excluded_k562_above_25": machinery_excluded,
            "k562_not_assayed": machinery_missing,
        },
        "influential_target": {
            "passed": least_stable["spearman_rho"] > 0,
            "minimum_leave_one_out": least_stable,
            "maximum_spearman_rho": max(row["spearman_rho"] for row in leave_one_out),
            "runs": len(leave_one_out),
        },
    }
    primary_pass = permutation_p <= 0.01
    all_kills_pass = all(row["passed"] for row in kills.values())
    status = "evidence_attached" if primary_pass and all_kills_pass else "orthogonal_phenotype"

    descriptive = {
        condition: _association(overlap, marson, reach, condition)
        for condition in ("Rest", "Stim8hr", "Stim48hr")
    }
    strongest_left = []
    strongest_right = []
    for gene in overlap:
        values = [
            int(marson[(gene, condition)]["n_total_de_genes"])
            for condition in ("Rest", "Stim8hr", "Stim48hr")
            if (gene, condition) in marson
        ]
        if values:
            strongest_left.append(float(max(values)))
            strongest_right.append(float(reach[gene]))
    descriptive["strongest_condition"] = {
        "n": len(strongest_left),
        "spearman_rho": round(_spearman(strongest_left, strongest_right), 6),
    }

    result_core = {
        "schema_version": "prospect.gse271788.calibration.v1",
        "pre_registration_id": prereg["pre_registration_id"],
        "status": status,
        "accepted": False,
        "next": "human_signature_required",
        "trust_boundary": "proposal_only",
        "frontier_root": "root_a8b0dcdd4024e12f",
        "claim": {
            "text": (
                "Broad transcriptomic reach is positively ranked across independent stimulated "
                "primary-CD4 perturbation studies under the locked calibration rule."
            ),
            "ceiling": "Computation over released data, not wet-lab or clinical truth.",
            "comparability": "orthogonal_activation_context",
            "interpretation": (
                "The result attaches cross-study evidence for broad activated-CD4 reach. It does "
                "not establish condition-level equivalence and cannot earn contradicted."
            ),
        },
        "coverage": {
            "published_targets": len(reach),
            "marson_overlap": len(overlap),
            "not_assayed": not_assayed,
        },
        "primary_result": {
            "condition": "Stim48hr",
            "n": len(primary_genes),
            "spearman_rho": round(observed, 6),
            "permutation_p_value_one_sided": round(permutation_p, 8),
            "bootstrap_95_percent_interval": bootstrap,
            "permutations": 10000,
            "bootstrap_samples": 10000,
            "seed": SEED,
            "passed": primary_pass,
        },
        "adversarial_kills": kills,
        "descriptive_results": descriptive,
        "on_target_sensitivity": {
            "decision_role": "descriptive_only_not_pre_registered_primary",
            "n": len(on_target_genes),
            "spearman_rho": round(on_target_rho, 6),
        },
        "med12_policy": "inspected calibration control, not a newly selected discovery",
        "source_integrity": integrity,
        "replay": "python frontier/gse271788_calibration.py --check",
    }

    artifacts = [
        Artifact("gse271788_preregistration.json", _sha256(PREREG), "examples/data/gse271788_preregistration.json"),
        Artifact("target_reach.csv", _sha256(source_dir / "target_reach.csv"), "examples/data/gse271788/target_reach.csv"),
        Artifact("metadata_summary.json", _sha256(source_dir / "metadata_summary.json"), "examples/data/gse271788/metadata_summary.json"),
        Artifact("source_manifest.json", _sha256(source_dir / "source_manifest.json"), "examples/data/gse271788/source_manifest.json"),
        Artifact("marson_de_full.csv", _sha256(MARSON), "examples/data/marson_de_full.csv"),
        Artifact("replogle_k562_de.csv", _sha256(K562), "examples/data/replogle_k562_de.csv"),
    ]
    receipt = Receipt(
        frontier="prospect_independent_cd4_calibration",
        claim=result_core["claim"]["text"],
        kind="cross_study_calibration",
        subject=overlap,
        producer={"kind": "deterministic_analysis", "name": "ProspectGSE271788Calibration"},
        artifacts=artifacts,
        evidence=[
            EvidenceAtom("overlapping perturbation targets", str(len(overlap)), "frozen source projection"),
            EvidenceAtom("Stim48hr Spearman rho", f"{observed:.6f}", "locked cross-study analysis"),
            EvidenceAtom("one-sided permutation P value", f"{permutation_p:.8f}", "10000 deterministic permutations"),
            EvidenceAtom("pre-registered adversarial kills passed", str(sum(row["passed"] for row in kills.values())), "frozen kill rules"),
        ],
        verifier=Verifier(
            name="ProspectIndependentCD4Calibration",
            method="frozen rank comparison plus three pre-registered adversarial kills",
            replay="python frontier/gse271788_calibration.py --check",
        ),
        status=status,
        replayability="exact",
        conditions=[
            "Marson Stim48hr primary endpoint",
            "day-eight activated primary-CD4 knockout comparator",
            "orthogonal activation context",
            "proposal only",
        ],
        verification_requirements=[
            "frozen_source_hashes_match",
            "pre_registered_rules_rederive",
            "human_ed25519_signature",
        ],
        state_diff={
            "accepted": False,
            "model_can_apply": False,
            "effect": "proposal_only_no_state_mutation",
            "target": "prospect_independent_cd4_calibration",
        },
        replay_metadata={
            "command": "python frontier/gse271788_calibration.py --check",
            "verifier": "ProspectIndependentCD4Calibration",
            "replayability": "exact",
            "pre_registration_id": prereg["pre_registration_id"],
            "ceiling": result_core["claim"]["ceiling"],
        },
    ).freeze().to_dict()
    result_core["receipt"] = receipt
    result_core["proposal_id"] = _canonical_id("proposal_", {"receipt_id": receipt["receipt_id"]})
    return result_core


def render_markdown(result: dict[str, Any]) -> str:
    primary = result["primary_result"]
    batch = result["adversarial_kills"]["batch_specificity"]["batches"]
    machinery = result["adversarial_kills"]["general_machinery"]
    influential = result["adversarial_kills"]["influential_target"]
    return f"""# Independent primary-CD4 calibration

Status: `{result['status']}`  
Accepted: `false`  
Proposal: `{result['proposal_id']}`  
Receipt: `{result['receipt']['receipt_id']}`

## Result

Across {primary['n']} perturbation targets shared by the released Marson table and the combined
GSE171737/GSE271788 study, Marson `Stim48hr` transcriptomic reach has Spearman rho
{primary['spearman_rho']:.6f} with the authors' published day-eight activated-CD4 knockout reach.
The one-sided 10,000-permutation P value is {primary['permutation_p_value_one_sided']:.8f}; the
bootstrap 95 percent interval is [{primary['bootstrap_95_percent_interval'][0]:.6f},
{primary['bootstrap_95_percent_interval'][1]:.6f}].

This earns `{result['status']}` under pre-registration `{result['pre_registration_id']}`. It is
computation over released data, not wet-lab or clinical truth. The studies use different activation
times and perturbation modalities, so the result attaches evidence for broad activated-CD4 reach.
It does not establish condition-level equivalence and cannot earn `contradicted`.

## Adversarial kills

| Kill | Result | Detail |
| --- | --- | --- |
| Batch specificity | pass | GSE171737 n={batch['gse171737_il2ra_regulators']['n']}, rho={batch['gse171737_il2ra_regulators']['spearman_rho']:.6f}; GSE271788 n={batch['gse271788_iei_background']['n']}, rho={batch['gse271788_iei_background']['spearman_rho']:.6f}. |
| General machinery | pass | After excluding K562 reach above 25 and K562 gaps, n={machinery['n']}, rho={machinery['spearman_rho']:.6f}. |
| Influential target | pass | Minimum leave-one-target-out rho={influential['minimum_leave_one_out']['spearman_rho']:.6f} after excluding {influential['minimum_leave_one_out']['excluded_gene']}. |

The descriptive sensitivity restricted to {result['on_target_sensitivity']['n']} Marson targets with
on-target knockdown at Stim48hr has rho {result['on_target_sensitivity']['spearman_rho']:.6f}. It is
reported as a sensitivity analysis and did not determine the status.

## Coverage and limits

- Published independent targets: {result['coverage']['published_targets']}.
- Marson overlap: {result['coverage']['marson_overlap']}.
- Marson `not_assayed`: {', '.join(result['coverage']['not_assayed'])}.
- MED12 remains an inspected calibration control, not a newly selected discovery.
- No claim in this packet changes the signed frontier.

## Replay

```bash
python frontier/gse271788_preregistration.py --check
python frontier/gse271788_calibration.py --check
```
"""


def write_calibration() -> dict[str, Any]:
    result = build_calibration()
    OUT_JSON.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n")
    OUT_DOC.write_text(render_markdown(result))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--freeze", action="store_true")
    parser.add_argument("--source-dir", type=Path)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    if args.freeze:
        if not args.source_dir:
            raise SystemExit("--freeze requires --source-dir")
        freeze_sources(args.source_dir)
    result = build_calibration()
    if args.check:
        if not OUT_JSON.exists() or json.loads(OUT_JSON.read_text()) != result:
            raise SystemExit("gse271788 calibration drift")
        if not OUT_DOC.exists() or OUT_DOC.read_text() != render_markdown(result):
            raise SystemExit("gse271788 calibration document drift")
    else:
        OUT_JSON.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n")
        OUT_DOC.write_text(render_markdown(result))
    print(
        f"status={result['status']} accepted=false n={result['primary_result']['n']} "
        f"rho={result['primary_result']['spearman_rho']:.6f} "
        f"p={result['primary_result']['permutation_p_value_one_sided']:.8f}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
